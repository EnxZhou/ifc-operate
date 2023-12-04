import csv

from OCC.Core import BRepBuilderAPI, BRepGProp, TopoDS

from OCC.Core.GProp import GProp_GProps
from OCC.Core.BRepGProp import brepgprop_VolumeProperties, brepgprop_LinearProperties
import FaceUtil
from OCC.Extend.DataExchange import read_step_file, read_step_file_with_names_colors
from OCC.Extend.TopologyUtils import is_face, is_edge, TopologyExplorer

import openpyxl


# 体特征
class SolidFeature:
    mass = 0.0
    centreOfMassX = 0.0
    centreOfMassY = 0.0
    centreOfMassZ = 0.0
    # 表面积
    surfaceArea = 0.0
    # 最大面特征
    maxFaceFeature = FaceUtil.FaceFeature
    # 最小面特征
    minFaceFeature = FaceUtil.FaceFeature
    # 面积平均值
    faceMassAverage = 0.0
    # 面积方差
    faceMassVariance = 0.0
    # 边数量
    edgeCount = 0
    # 边长总和
    edgeLenSum = 0.0

    def to_dict(self):
        return {
            "mass": self.mass,
            "centreOfMassX": self.centreOfMassX,
            "centreOfMassY": self.centreOfMassY,
            "centreOfMassZ": self.centreOfMassZ,
            "surfaceArea": self.surfaceArea,
            "faceMassAverage": self.faceMassAverage,
            "faceMassVariance": self.faceMassVariance,
            "edgeCount": self.edgeCount,
            "edgeLenSum": self.edgeLenSum
        }


# solid节点的特征
class SolidNodeFeature:
    mass = 0.0
    centreOfMassX = 0.0
    centreOfMassY = 0.0
    centreOfMassZ = 0.0
    faceCount = 0

    def to_dict(self):
        return {
            "mass": self.mass,
            "centreOfMassX": self.centreOfMassX,
            "centreOfMassY": self.centreOfMassY,
            "centreOfMassZ": self.centreOfMassZ,
            "faceCount": self.faceCount,
        }

    def to_list(self):
        return [self.mass,
                self.centreOfMassX,
                self.centreOfMassY,
                self.centreOfMassZ,
                self.faceCount]


# SolidNode代表将Solid作为graph的节点（node）处理
# 节点属性包含节点名称即节点特征
class SolidNode:
    name = ""
    feature = SolidNodeFeature

    def __init__(self, name, feature):
        self.name = name
        self.feature = feature


# 读取判断好dist的文件，转换为SolidNode
def solid_node_from_csv(file_name):
    # 读取xlsx文件中名为“solid”的sheet表
    wb = openpyxl.load_workbook(file_name, read_only=True)
    ws = wb['solid']

    # 获取字段名所在的行号
    header_row = 1
    header = []
    for cell in ws[header_row]:
        header.append(cell.value)

    # 获取每个字段对应的列号
    mass_col = header.index('mass')
    centreOfMassX_col = header.index('centreOfMassX')
    centreOfMassY_col = header.index('centreOfMassY')
    centreOfMassZ_col = header.index('centreOfMassZ')
    faceCount_col = header.index('faceCount')

    # 遍历sheet表中的数据，并将其写入SolidNode对象
    solid_nodes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[header.index('name')]
        feature = SolidNodeFeature()
        feature.mass = row[mass_col]
        feature.centreOfMassX = row[centreOfMassX_col]
        feature.centreOfMassY = row[centreOfMassY_col]
        feature.centreOfMassZ = row[centreOfMassZ_col]
        feature.faceCount = row[faceCount_col]
        solid_node = SolidNode(name, feature)
        solid_nodes.append(solid_node)
    return solid_nodes


def get_solid_node(_solid):
    props = GProp_GProps()
    # 创建计算几何属性对象
    brepgprop_VolumeProperties(_solid, props)
    # 计算重量和重心
    mass = props.Mass()
    centre_of_mass = props.CentreOfMass()
    # 获取面个数
    face_count = FaceUtil.get_face_count_of_solid(_solid)

    # 创建 SolidNodeFeature 对象
    feature = SolidNodeFeature()
    feature.mass = mass
    feature.centreOfMassX = centre_of_mass.X()
    feature.centreOfMassY = centre_of_mass.Y()
    feature.centreOfMassZ = centre_of_mass.Z()
    feature.faceCount = face_count

    return feature


def get_solids_from_step(file_path):
    stpshp = read_step_file(file_path)

    # step文件生成的shape
    stpshp2 = TopologyExplorer(stpshp)
    # 获取文件中的体
    solidList = stpshp2.solids()
    return solidList


def get_solid_feature(_solid):
    props = GProp_GProps()
    # 创建计算几何属性对象
    brepgprop_VolumeProperties(_solid, props)
    # 计算重量和重心
    mass = props.Mass()
    centre_of_mass = props.CentreOfMass()
    # 获取表面积
    face_list = FaceUtil.get_faces_from_solid(_solid)
    surface_area = 0.0

    # 创建最大面和最小面的变量
    max_face = None
    max_area = 0.0
    min_face = None
    min_area = float('inf')
    face_masses = []

    # 遍历 Solid 中的每个面，获取面积最大和最小的面
    for face in face_list:
        area = FaceUtil.get_face_props(face).mass
        face_masses.append(area)
        if area > max_area:
            max_area = area
            max_face = face
        if area < min_area:
            min_area = area
            min_face = face

    # 计算面积平均值和方差
    surface_area = sum(face_masses)
    face_mass_average = sum(face_masses) / len(face_masses)
    face_mass_variance = sum([(m - face_mass_average) ** 2 \
                              for m in face_masses]) / len(face_masses)

    # 创建 SolidFeature 对象
    feature = SolidFeature()
    feature.mass = mass
    feature.centreOfMassX = centre_of_mass.X()
    feature.centreOfMassY = centre_of_mass.Y()
    feature.centreOfMassZ = centre_of_mass.Z()
    feature.surfaceArea = surface_area
    feature.maxFaceFeature = FaceUtil.get_face_props(max_face)
    feature.minFaceFeature = FaceUtil.get_face_props(min_face)
    feature.faceMassAverage = face_mass_average
    feature.faceMassVariance = face_mass_variance
    feature.edgeLenSum = get_solid_edge_len_sum(_solid)
    feature.edgeCount = get_solid_edge_count(_solid)

    return feature


# 获取边长总和
def get_solid_edge_len_sum(_solid):
    t = TopologyExplorer(_solid)
    props = GProp_GProps()
    edgeLenSum = 0.0
    for edge in t.edges():
        brepgprop_LinearProperties(edge, props)
        edge_len = props.Mass()
        edgeLenSum += edge_len
    return edgeLenSum


# 获取边数量
def get_solid_edge_count(_solid):
    t = TopologyExplorer(_solid)
    edgeCount = t.number_of_edges()
    return edgeCount


def hasSharedFace(_solid1, _solid2):
    # 定义一个空的共享面列表
    shared_faces = []

    # 遍历 solid1 的所有面
    solid1FaceList = FaceUtil.get_faces_from_solid(_solid1)
    solid2FaceList = FaceUtil.get_faces_from_solid(_solid2)
    for face1 in solid1FaceList:
        # 遍历 solid2 的所有面
        for face2 in solid2FaceList:
            # 判断 face1 和 face2 是否相同
            if face1.IsSame(face2):
                # 如果相同，说明这是一个共享面
                shared_faces.append(face1)

    # 如果共享面列表不为空，则说明 solid1 和 solid2 有共享面
    if shared_faces:
        print("Solid1 and Solid2 share the following faces:")
        for face in shared_faces:
            print(face)
    else:
        print("Solid1 and Solid2 do not share any faces.")
