import numpy as np
import openpyxl
import networkx as nx
import pandas as pd


# ��������ڵ�
class MainSubNode:
    id = 0
    guid = ''
    node_class = ''
    centre_of_mass_x = 0.0
    centre_of_mass_y = 0.0
    centre_of_mass_z = 0.0
    max_face_axis_location_x = 0.0
    max_face_axis_location_y = 0.0
    max_face_axis_location_z = 0.0
    max_face_axis_direct_x = 0.0
    max_face_axis_direct_y = 0.0
    max_face_axis_direct_z = 0.0

    def __init__(self,
                 id,
                 guid,
                 node_class,
                 centre_of_mass_x,
                 centre_of_mass_y,
                 centre_of_mass_z,
                 max_face_axis_location_x,
                 max_face_axis_location_y,
                 max_face_axis_location_z,
                 max_face_axis_direct_x,
                 max_face_axis_direct_y,
                 max_face_axis_direct_z):
        self.id = id
        self.guid = guid
        self.node_class = node_class
        self.centre_of_mass_x = centre_of_mass_x
        self.centre_of_mass_y = centre_of_mass_y
        self.centre_of_mass_z = centre_of_mass_z
        self.max_face_axis_location_x = max_face_axis_location_x
        self.max_face_axis_location_y = max_face_axis_location_y
        self.max_face_axis_location_z = max_face_axis_location_z
        self.max_face_axis_direct_x = max_face_axis_direct_x
        self.max_face_axis_direct_y = max_face_axis_direct_y
        self.max_face_axis_direct_z = max_face_axis_direct_z


# ��ȡ���ɵ�solid����xlsx�ļ����γ�SolidFeature�б�
def read_main_sub_node_from_xlsx(file_name):
    # ��ȡxlsx�ļ�����Ϊ��node����sheet��
    wb = openpyxl.load_workbook(file_name, read_only=True)
    ws = wb['solid']

    # ��ȡ�ֶ������ڵ��к�
    header_row = 1
    header = []
    for cell in ws[header_row]:
        header.append(cell.value)

    # ��ȡÿ���ֶζ�Ӧ���к�
    node_id_col = header.index('id')
    guid_col = header.index('guid')
    node_class_col = header.index('class')
    centre_of_mass_x_col = header.index('centreOfMassX')
    centre_of_mass_y_col = header.index('centreOfMassY')
    centre_of_mass_z_col = header.index('centreOfMassZ')
    max_face_axis_direct_x_col = header.index('maxFaceAxisDirectX')
    max_face_axis_direct_y_col = header.index('maxFaceAxisDirectY')
    max_face_axis_direct_z_col = header.index('maxFaceAxisDirectZ')
    max_face_axis_location_x_col = header.index('maxFaceAxisLocationX')
    max_face_axis_location_y_col = header.index('maxFaceAxisLocationY')
    max_face_axis_location_z_col = header.index('maxFaceAxisLocationZ')

    # ����sheet���е����ݣ�������д��SolidNode����
    main_sub_nodes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        node_id = row[node_id_col]
        guid = row[guid_col]
        node_class = row[node_class_col]
        centre_of_mass_x = row[centre_of_mass_x_col]
        centre_of_mass_y = row[centre_of_mass_y_col]
        centre_of_mass_z = row[centre_of_mass_z_col]
        max_face_axis_location_x = row[max_face_axis_location_x_col]
        max_face_axis_location_y = row[max_face_axis_location_y_col]
        max_face_axis_location_z = row[max_face_axis_location_z_col]
        max_face_axis_direct_x = row[max_face_axis_direct_x_col]
        max_face_axis_direct_y = row[max_face_axis_direct_y_col]
        max_face_axis_direct_z = row[max_face_axis_direct_z_col]
        main_sub_node = MainSubNode(node_id,
                                    guid,
                                    node_class,
                                    centre_of_mass_x,
                                    centre_of_mass_y,
                                    centre_of_mass_z,
                                    max_face_axis_location_x,
                                    max_face_axis_location_y,
                                    max_face_axis_location_z,
                                    max_face_axis_direct_x,
                                    max_face_axis_direct_y,
                                    max_face_axis_direct_z)
        main_sub_nodes.append(main_sub_node)
    return main_sub_nodes


class Face:
    def __init__(self, normal_vector, point):
        self.normal_vector = np.array(normal_vector)
        self.point = np.array(point)

    def __str__(self):
        return f"Normal Vector: {self.normal_vector}, Point: {self.point}"

    def shortest_distance_to_point(self, test_point):
        np_test_point = np.array(test_point)
        # Calculate the vector from the plane's point to the test point
        vector_to_point = np_test_point - self.point

        # Calculate the projection of the vector to the point onto the normal vector
        projection = np.dot(vector_to_point, self.normal_vector) / np.linalg.norm(self.normal_vector)

        # Calculate the shortest distance
        distance = abs(projection)
        return distance


# �ж����������Ƿ���
def are_vectors_collinear(x1, y1, z1, x2, y2, z2, tolerance=1e-6):
    # Calculate the magnitudes of the vectors
    magnitude1 = (x1 ** 2 + y1 ** 2 + z1 ** 2) ** 0.5
    magnitude2 = (x2 ** 2 + y2 ** 2 + z2 ** 2) ** 0.5

    # Check if the magnitudes are approximately equal (within the tolerance)
    magnitudes_equal = abs(magnitude1 - magnitude2) < tolerance

    # Check if the cross product of the two vectors is zero
    cross_product_x = y1 * z2 - z1 * y2
    cross_product_y = z1 * x2 - x1 * z2
    cross_product_z = x1 * y2 - y1 * x2

    cross_product_magnitude = (cross_product_x ** 2 + cross_product_y ** 2 + cross_product_z ** 2) ** 0.5

    # Check if the cross product magnitude is approximately zero (within the tolerance)
    cross_product_near_zero = cross_product_magnitude < tolerance

    # Check if both magnitude and direction are within tolerance
    return magnitudes_equal and cross_product_near_zero


# �ϲ���main node
# ������Ϊһ���嵥Ԫ
class CombineMainNode:
    name = ""
    main_node_list = []


def extract_collinear_connected_nodes(graph: nx.Graph):
    collinear_connected_nodes = []
    for node1, node2 in graph.edges():
        try:
            normals1 = (graph.nodes[node1]['max_face_axis_direct_x'],
                        graph.nodes[node1]['max_face_axis_direct_y'],
                        graph.nodes[node1]['max_face_axis_direct_z'])
            normals2 = (graph.nodes[node2]['max_face_axis_direct_x'],
                        graph.nodes[node2]['max_face_axis_direct_y'],
                        graph.nodes[node2]['max_face_axis_direct_z'])
            node1_class = graph.nodes[node1]['node_class']
            node2_class = graph.nodes[node2]['node_class']
            if not (node1_class == 'main' and node2_class == 'main'):
                continue
            if are_vectors_collinear(*normals1, *normals2):
                collinear_connected_nodes.append((node1, node2))
        except KeyError as e:
            print(f"KeyError: {e} occurred for nodes {node1} and {node2}")
    return collinear_connected_nodes


# ������ȱ������������main node�ϲ�
def dfs_main_coplanar(graph, node, visited, current_set):
    visited[node] = True
    current_set.append(node)

    for neighbor in graph.neighbors(node):
        edge_feature = graph[node][neighbor].get('coplanar', False)
        if not visited[neighbor] and edge_feature:
            dfs_main_coplanar(graph, neighbor, visited, current_set)


# �ж�һ���б� a �Ƿ��� coplanar_main_set �е�һ���б���Ӽ�
def is_subset_of_coplanar_set(a, coplanar_main_set):
    for subset in coplanar_main_set:
        if set(a).issubset(subset):
            return True
    return False


def is_majority_subset_of_coplanar_set(a, coplanar_main_set):
    subset_counts = {}  # Dictionary to store counts of lists in coplanar_main_set
    for subset in coplanar_main_set:
        count = sum(1 for lst in a if lst in subset)
        subset_counts[tuple(subset)] = count

    max_count_subset = max(subset_counts, key=subset_counts.get)
    max_count = subset_counts[max_count_subset]
    return max_count_subset, max_count


def handle_nearest_main_node_more_than_one(node: str,
                                           nearest_main_nodes: list,
                                           graph: nx.Graph,
                                           coplanar_main_sets: []):
    max_axis_direct_x_tolerance = 0.001
    if len(nearest_main_nodes) == 0:
        raise ValueError("No nearest main node")
    # ����������main node�ж��
    if len(nearest_main_nodes) > 1:
        max_face_axis_direct_x = graph.nodes[node]['max_face_axis_direct_x']
        # Check if all nearest_main_nodes belong to the same main_node_group
        # ��������main node����ͬһ��main group����ֱ�ӹ�������һ��group
        same_group = is_subset_of_coplanar_set(nearest_main_nodes, coplanar_main_sets)
        if same_group:
            selected_main_node = nearest_main_nodes[0]
            return selected_main_node

        max_count_subset, max_count = is_majority_subset_of_coplanar_set(nearest_main_nodes,
                                                                         coplanar_main_sets)
        if max_count >= 0.33 * len(nearest_main_nodes):
            selected_main_node = max_count_subset[0]
            return selected_main_node

        else:
            for main_node in nearest_main_nodes:
                main_axis_direct_x = graph.nodes[main_node]['max_face_axis_direct_x']
                if abs(main_axis_direct_x - max_face_axis_direct_x) <= max_axis_direct_x_tolerance:
                    selected_main_node = main_node
                    return selected_main_node
    # ���ֻ��һ�����main node���Ͳ���Ҫ����һ���ж���
    else:
        selected_main_node = nearest_main_nodes[0]
        return selected_main_node


# ����һ��node������һ��node�����ľ���
def distance_to_node_max_face(node: str, main_node: str, graph: nx.Graph):
    node_centre_x = graph.nodes[node]['centre_of_mass_x']
    node_centre_y = graph.nodes[node]['centre_of_mass_y']
    node_centre_z = graph.nodes[node]['centre_of_mass_z']
    node_centre = [node_centre_x, node_centre_y, node_centre_z]

    max_face_axis_location_x = graph.nodes[main_node]['max_face_axis_location_x']
    max_face_axis_location_y = graph.nodes[main_node]['max_face_axis_location_y']
    max_face_axis_location_z = graph.nodes[main_node]['max_face_axis_location_z']
    max_face_axis_direct_x = graph.nodes[main_node]['max_face_axis_direct_x']
    max_face_axis_direct_y = graph.nodes[main_node]['max_face_axis_direct_y']
    max_face_axis_direct_z = graph.nodes[main_node]['max_face_axis_direct_z']
    normal_vector = [max_face_axis_direct_x, max_face_axis_direct_y, max_face_axis_direct_z]
    normal_location = [max_face_axis_location_x, max_face_axis_location_y, max_face_axis_location_z]
    max_face = Face(normal_vector, normal_location)
    distance = max_face.shortest_distance_to_point(node_centre)
    return distance


def handle_nearest_main_node_more_than_one2(node: str,
                                            nearest_main_nodes: list,
                                            graph: nx.Graph,
                                            coplanar_main_sets: []):
    selected_main_node = None
    if len(nearest_main_nodes) == 0:
        raise ValueError("No nearest main node")
    # ����������main node�ж��
    if len(nearest_main_nodes) > 1:

        # Check if all nearest_main_nodes belong to the same main_node_group
        # ��������main node����ͬһ��main group����ֱ�ӹ�������һ��group
        same_group = is_subset_of_coplanar_set(nearest_main_nodes, coplanar_main_sets)
        if same_group:
            selected_main_node = nearest_main_nodes[0]
            return selected_main_node

        node_centre_x = graph.nodes[node]['centre_of_mass_x']
        node_centre_y = graph.nodes[node]['centre_of_mass_y']
        node_centre_z = graph.nodes[node]['centre_of_mass_z']
        node_centre = [node_centre_x, node_centre_y, node_centre_z]
        min_distance = float('inf')
        for main_node in nearest_main_nodes:
            max_face_axis_location_x = graph.nodes[main_node]['max_face_axis_location_x']
            max_face_axis_location_y = graph.nodes[main_node]['max_face_axis_location_y']
            max_face_axis_location_z = graph.nodes[main_node]['max_face_axis_location_z']
            max_face_axis_direct_x = graph.nodes[main_node]['max_face_axis_direct_x']
            max_face_axis_direct_y = graph.nodes[main_node]['max_face_axis_direct_y']
            max_face_axis_direct_z = graph.nodes[main_node]['max_face_axis_direct_z']
            normal_vector = [max_face_axis_direct_x, max_face_axis_direct_y, max_face_axis_direct_z]
            normal_location = [max_face_axis_location_x, max_face_axis_location_y, max_face_axis_location_z]
            max_face = Face(normal_vector, normal_location)
            distance = max_face.shortest_distance_to_point(node_centre)
            if distance < min_distance:
                min_distance = distance
                selected_main_node = main_node

        return selected_main_node
    # ���ֻ��һ�����main node���Ͳ���Ҫ����һ���ж���
    else:
        selected_main_node = nearest_main_nodes[0]
        return selected_main_node


def handle_nearest_main_node(node: str,
                             nearest_main_nodes: list,
                             graph: nx.Graph):
    selected_main_node = None
    if len(nearest_main_nodes) == 0:
        raise ValueError("No nearest main node")
    # ���ֻ��һ�����main node���Ͳ���Ҫ����һ���ж���
    if len(nearest_main_nodes) == 1:
        return nearest_main_nodes[0]
    # ����������main node�ж��
    else:
        min_distance = float('inf')
        for main_node in nearest_main_nodes:
            distance = nx.shortest_path_length(graph, source=node, target=main_node)
            if distance < min_distance:
                min_distance = distance
                selected_main_node = main_node

        return selected_main_node


def find_short_paths_within_distance(graph, source, target, max_distance, exclude_nodes):
    # ���������ڵ�֮������·��
    # if source == '614' and target=='38':
    #     print(1)
    all_paths = list(nx.all_shortest_paths(graph, source=source, target=target, weight='weight'))
    for path in all_paths:
        if len(path) <= max_distance:
            # ���·�����Ƿ������Ҫ�ų��Ľڵ�
            if any(node in exclude_nodes for node in path):
                return None
            return path

    return None


def get_exclude_nodes(all_node_dict: dict, remove_node):
    tmp_dict = all_node_dict.copy()
    del tmp_dict[remove_node]
    remain_keys_list = list(tmp_dict.keys())
    return remain_keys_list


def extract_sub_rely_main2(graph: nx.Graph, coplanar_main_sets):
    sub_to_main_group_dict = {}
    main_node_dict = {}

    for i, main_node_group in enumerate(coplanar_main_sets):
        for main_node in main_node_group:
            main_node_dict[main_node] = i

    for node in graph.nodes:
        node_class = graph.nodes[node]['node_class']

        # Only parse the sub node
        if not node_class == 'sub':
            continue

        nearest_main_nodes = []
        min_distance = float('inf')
        for main_node_group in coplanar_main_sets:
            for main_node in main_node_group:
                # �п�������node֮��û��path
                if nx.has_path(graph, source=node, target=main_node):
                    distance = nx.shortest_path_length(graph, source=node, target=main_node)
                    if distance < min_distance:
                        min_distance = distance
                        nearest_main_nodes = [main_node]
                    elif distance == min_distance:
                        nearest_main_nodes.append(main_node)
                    nearest_main_nodes.append(main_node)

        # if node == '475':
        #     print(1)
        selected_main_node = handle_nearest_main_node_more_than_one2(node,
                                                                     nearest_main_nodes,
                                                                     graph,
                                                                     coplanar_main_sets)

        if selected_main_node is None:
            print(f"node:{node} not have main node")

        if selected_main_node is not None:
            sub_to_main_group_dict[node] = main_node_dict[selected_main_node]

    return sub_to_main_group_dict


def extract_sub_rely_main3(graph: nx.Graph, coplanar_main_sets):
    sub_to_main_group_dict = {}
    main_node_dict = {}

    for i, main_node_group in enumerate(coplanar_main_sets):
        for main_node in main_node_group:
            main_node_dict[main_node] = i

    for node in graph.nodes:
        node_class = graph.nodes[node]['node_class']

        # Only parse the sub node
        if not node_class == 'sub':
            continue

        nearest_main_nodes = []
        distance_dict = {}
        # �ݲ��ǹؼ����ݲ��������main node���ܱ�������ȥ
        # �ݲС�������ų�һЩ����
        # xm-60
        # sc-20

        tolerance = 20
        # ����node���ľ�����Щmain node�����������
        for main_node_group in coplanar_main_sets:
            for main_node in main_node_group:
                distance = distance_to_node_max_face(node, main_node, graph)
                distance_dict[main_node]=distance

        sorted_dict = dict(sorted(distance_dict.items(), key=lambda item: item[1]))
        min_distance = next(iter(sorted_dict.items()))[1]
        for main_node,distance in sorted_dict.items():
            if distance - min_distance <=tolerance:
                nearest_main_nodes.append(main_node)

        selected_main_node = handle_nearest_main_node(node,
                                                      nearest_main_nodes,
                                                      graph)

        if selected_main_node is None:
            print(f"node:{node} not have main node")
        else:
            try:
                sub_to_main_group_dict[node] = main_node_dict[selected_main_node]
            except Exception as e:
                print("An error occurred:", e)
                print("node:", node)
                print("selected_main_node:", selected_main_node)
                print("main_node_dict:", main_node_dict)
                print("sub_to_main_group_dict:", sub_to_main_group_dict)

    return sub_to_main_group_dict


# ��������main node��������ڵ�����main node���棬�ͽ������ͬһ��������
# ���صĽ������һ��main node�ļ��ϵ�list
# list�е�ÿһ�����һ�����ϣ��ü����е�main node��Ӧ����һ���嵥Ԫ
# �ϲ�main node�Ĺ���
# 1.main node֮���б�ֱ������
# 2.����淨�ߣ�maxFaceAxis��������
def extract_main_node(graph):
    # set all node as not visited
    visited = {node: False for node in graph.nodes()}
    node_sets = []

    for node in graph.nodes():
        node_class = graph.nodes[node]['node_class']
        # just parse the main node
        if not node_class == 'main':
            continue
        if not visited[node]:
            current_set = []
            dfs_main_coplanar(graph, node, visited, current_set)
            if current_set:
                node_sets.append(current_set)

    return node_sets


# ��mian node��sub node�ϲ����γ������嵥Ԫ
def combine_node(graph: nx.Graph):
    collinear_edge = extract_collinear_connected_nodes(graph)
    for node1, node2 in collinear_edge:
        graph[node1][node2]['coplanar'] = True

    # ��ȡ���������Ľڵ㼯��
    coplanar_main_sets = extract_main_node(graph)

    sub_to_main_group_dict = extract_sub_rely_main3(graph, coplanar_main_sets)

    combine_node_dict = {}
    # ��main node�Ľ���Ž�ȥ��key�������
    for i, main_node_group in enumerate(coplanar_main_sets):
        combine_node_dict[i] = main_node_group

    # ��sub node�Ľ���Ž�ȥ��key��value
    for k, v in sub_to_main_group_dict.items():
        combine_node_dict[v].append(k)

    return combine_node_dict
