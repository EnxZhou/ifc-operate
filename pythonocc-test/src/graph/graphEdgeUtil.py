import openpyxl


class EdgeOfFaceNode:
    node1 = ""
    node2 = ""

    def __init__(self, node1: str, node2: str):
        self.node1 = node1
        self.node2 = node2


# ��ȡ�жϺ�dist���ļ���ת��Ϊͼ�ṹ��edge
def dist_edge_from_xlsx(file_name):
    # ��ȡxlsx�ļ�����Ϊ��dist����sheet��
    wb = openpyxl.load_workbook(file_name, read_only=True)
    ws = wb['dist']

    # ��ȡ�ֶ������ڵ��к�
    header_row = 1
    header = []
    for cell in ws[header_row]:
        header.append(cell.value)

    # ��ȡÿ���ֶζ�Ӧ���к�
    node1_col = header.index('node1')
    index1_col = header.index('index1')
    node2_col = header.index('node2')
    index2_col = header.index('index2')

    # ����sheet���е����ݣ�������д��SolidNode����
    edge_nodes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        node1 = row[node1_col] + '_' + str(row[index1_col])
        node2 = row[node2_col] + '_' + str(row[index2_col])
        edge_node = EdgeOfFaceNode(node1, node2)
        edge_nodes.append(edge_node)
    return edge_nodes


# ��ȡ�ڵ�edge�ļ����γ�ͼ�ı�����
def read_edge_from_xlsx(file_name):
    # ��ȡxlsx�ļ�����Ϊ��edge����sheet��
    wb = openpyxl.load_workbook(file_name, read_only=True)
    ws = wb['edge']

    # ��ȡ�ֶ������ڵ��к�
    header_row = 1
    header = []
    for cell in ws[header_row]:
        header.append(cell.value)

    # ��ȡÿ���ֶζ�Ӧ���к�
    node1_col = header.index('Src')
    node2_col = header.index('Dst')

    # ����sheet���е����ݣ�������д��SolidNode����
    edge_nodes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        node1 = row[node1_col]
        node2 = row[node2_col]
        edge_node = EdgeOfFaceNode(node1, node2)
        edge_nodes.append(edge_node)
    return edge_nodes
