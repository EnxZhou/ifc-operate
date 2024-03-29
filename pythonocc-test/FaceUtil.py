import math

from OCC.Core.BRepAdaptor import BRepAdaptor_Surface
from OCC.Core.GProp import GProp_GProps
from OCC.Core.TopoDS import TopoDS_Face, TopoDS_Shape
from OCC.Core.BRepGProp import brepgprop_LinearProperties, brepgprop_SurfaceProperties, brepgprop_VolumeProperties
from OCC.Core.GeomAbs import (
    GeomAbs_C0,
    GeomAbs_Plane,
    GeomAbs_Cylinder,
    GeomAbs_Cone,
    GeomAbs_Sphere,
    GeomAbs_Torus,
    GeomAbs_BezierSurface,
    GeomAbs_BSplineSurface,
    GeomAbs_SurfaceOfRevolution,
    GeomAbs_SurfaceOfExtrusion,
    GeomAbs_OffsetSurface,
    GeomAbs_OtherSurface,
)
from OCC.Core.gp import gp_Pnt, gp_Dir
from OCC.Extend.TopologyUtils import TopologyExplorer
from OCC.Core.BRepTools import breptools_UVBounds

import openpyxl

import EdgeUtil


# 面的特征
class FaceFeature:
    kind = ""
    mass = 0.0
    centreOfMassX = 0.0
    centreOfMassY = 0.0
    centreOfMassZ = 0.0
    axisLocationX = 0.0
    axisLocationY = 0.0
    axisLocationZ = 0.0
    axisDirectX = 0.0
    axisDirectY = 0.0
    axisDirectZ = 0.0
    # 周长
    perimeter = 0.0
    # 最长边特征
    maxEdgeFeature = EdgeUtil.EdgeFeature
    # 最短边特征
    minEdgeFeature = EdgeUtil.EdgeFeature
    edgeLengthAverage = 0.0
    # 边长方差
    edgeLengthVariance = 0.0

    def __init__(self):
        self.kind = ""
        self.mass = 0.0
        self.centreOfMassX = 0.0
        self.centreOfMassY = 0.0
        self.centreOfMassZ = 0.0
        self.axisLocationX = 0.0
        self.axisLocationY = 0.0
        self.axisLocationZ = 0.0
        self.axisDirectX = 0.0
        self.axisDirectY = 0.0
        self.axisDirectZ = 0.0
        self.perimeter = 0.0
        self.maxEdgeFeature = EdgeUtil.EdgeFeature()
        self.minEdgeFeature = EdgeUtil.EdgeFeature()
        self.edgeLengthAverage = 0.0
        self.edgeLengthVariance = 0.0

    def __str__(self):
        return f"FaceFeature({self.kind}," \
               f" {self.mass}," \
               f" ({self.centreOfMassX}," \
               f" {self.centreOfMassY}," \
               f" {self.centreOfMassZ})," \
               f" ({self.axisLocationX}," \
               f" {self.axisLocationY}," \
               f" {self.axisLocationZ})," \
               f" ({self.axisDirectX}," \
               f" {self.axisDirectY}," \
               f" {self.axisDirectZ})," \
               f" {self.perimeter}," \
               f" {self.maxEdgeFeature}," \
               f" {self.minEdgeFeature}," \
               f" {self.edgeLengthAverage}," \
               f" {self.edgeLengthVariance})"

    # def toStr(self):
    #     return f"kind: {self.kind}, mass: {self.mass}, centerOfMassX: {self.centreOfMassX}, " \
    #            f"centerOfMassY: {self.centreOfMassY}, centerOfMassZ: {self.centreOfMassZ}, " \
    #            f"axisLocationX: {self.axisLocationX}, axisLocationY: {self.axisLocationY}, " \
    #            f"axisLocationZ: {self.axisLocationZ}, axisDirectX: {self.axisDirectX}, " \
    #            f"axisDirectY: {self.axisDirectY}, axisDirectZ: {self.axisDirectZ}, perimeter: {self.perimeter}, " \
    #            f"maxEdgeFeature: {self.maxEdgeFeature}, minEdgeFeature: {self.minEdgeFeature}, " \
    #            f"edgeLengthAverage: {self.edgeLengthAverage}, edgeLengthVariance: {self.edgeLengthVariance}"


# 面节点的特征
class FaceNodeFeature:
    kind = ""
    mass = 0.0
    centreOfMassX = 0.0
    centreOfMassY = 0.0
    centreOfMassZ = 0.0
    axisLocationX = 0.0
    axisLocationY = 0.0
    axisLocationZ = 0.0
    axisDirectX = 0.0
    axisDirectY = 0.0
    axisDirectZ = 0.0
    edgeCount = 0

    def to_dict(self):
        return {
            "mass": self.mass,
            "centreOfMassX": self.centreOfMassX,
            "centreOfMassY": self.centreOfMassY,
            "centreOfMassZ": self.centreOfMassZ,
            "axisLocationX": self.axisLocationX,
            "axisLocationY": self.axisLocationY,
            "axisLocationZ": self.axisLocationZ,
            "axisDirectX": self.axisDirectX,
            "axisDirectY": self.axisDirectY,
            "axisDirectZ": self.axisDirectZ,
            "edgeCount": self.edgeCount,
        }

    def to_list(self):
        return [self.mass,
                self.centreOfMassX,
                self.centreOfMassY,
                self.centreOfMassZ,
                self.axisLocationX,
                self.axisLocationY,
                self.axisLocationZ,
                self.axisDirectX,
                self.axisDirectY,
                self.axisDirectZ,
                self.edgeCount]

    def list_marshal(self, data):
        self.mass = data[0]
        self.centreOfMassX = data[1]
        self.centreOfMassY = data[2]
        self.centreOfMassZ = data[3]
        self.axisLocationX = data[4]
        self.axisLocationY = data[5]
        self.axisLocationZ = data[6]
        self.axisDirectX = data[7]
        self.axisDirectY = data[8]
        self.axisDirectZ = data[9]
        self.edgeCount = data[10]


# 面节点
class FaceNode:
    topoDS_face = TopoDS_Face
    solidName = ""
    index = 0
    feature = FaceNodeFeature

    def __init__(self, topoDS_face, solidName, index, faceNodeFeature):
        self.topoDS_face = topoDS_face
        self.solidName = solidName
        self.index = index
        self.feature = faceNodeFeature


# 读取判断好dist的文件，转换为FaceNode
def face_node_from_xlsx(file_name):
    # 读取xlsx文件中名为“solid”的sheet表
    wb = openpyxl.load_workbook(file_name, read_only=True)
    ws = wb['face']

    # 获取字段名所在的行号
    header_row = 1
    header = []
    for cell in ws[header_row]:
        header.append(cell.value)

    # 获取每个字段对应的列号
    mass_col = header.index('mass')
    centreOfMassX_col = header.index('centreOfMassX')
    centreOfMassY_col = header.index('centreOfMassY')
    centreOfMassZ_col = header.index('centreOfMassZ')
    axisLocationX_col = header.index('axisLocationX')
    axisLocationY_col = header.index('axisLocationY')
    axisLocationZ_col = header.index('axisLocationZ')
    axisDirectX_col = header.index('axisDirectX')
    axisDirectY_col = header.index('axisDirectY')
    axisDirectZ_col = header.index('axisDirectZ')
    edgeCount_col = header.index('edgeCount')

    # 遍历sheet表中的数据，并将其写入SolidNode对象
    face_nodes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[header.index('name')]
        index = row[header.index('index')]
        feature = FaceNodeFeature()
        feature.mass = row[mass_col]
        feature.centreOfMassX = row[centreOfMassX_col]
        feature.centreOfMassY = row[centreOfMassY_col]
        feature.centreOfMassZ = row[centreOfMassZ_col]
        feature.axisLocationX = row[axisLocationX_col]
        feature.axisLocationY = row[axisLocationY_col]
        feature.axisLocationZ = row[axisLocationZ_col]
        feature.axisDirectX = row[axisDirectX_col]
        feature.axisDirectY = row[axisDirectY_col]
        feature.axisDirectZ = row[axisDirectZ_col]
        feature.edgeCount = row[edgeCount_col]
        face_node = FaceNode(topoDS_face=None,solidName=name, index=index, faceNodeFeature=feature)
        face_nodes.append(face_node)
    return face_nodes


# 获取面的属性
def get_face_props(_face: TopoDS_Face):
    props = GProp_GProps()
    brepgprop_SurfaceProperties(_face, props)

    faceFeatureReturn = FaceFeature()
    # 面积
    faceFeatureReturn.mass = props.Mass()
    centreOfMass = props.CentreOfMass()
    # 重心坐标
    faceFeatureReturn.centreOfMassX = centreOfMass.X()
    faceFeatureReturn.centreOfMassY = centreOfMass.Y()
    faceFeatureReturn.centreOfMassZ = centreOfMass.Z()
    # 创建最大边和最小边的变量
    maxEdgeFeature = None
    maxEdgeLen = 0.0
    minEdgeFeature = None
    minEdgeLen = float('inf')
    edge_len_list = []

    edge_list = EdgeUtil.get_edges_from_face(_face)

    # 遍历 face 中的每个edge，获取长度最大和最小的edge
    for edge in edge_list:
        edgeFeature = EdgeUtil.get_edge_props(edge)
        edgeLen = edgeFeature.length
        edge_len_list.append(edgeLen)
        if edgeLen > maxEdgeLen:
            maxEdgeLen = edgeLen
            maxEdgeFeature = edgeFeature
        if edgeLen < minEdgeLen:
            minEdgeLen = edgeLen
            minEdgeFeature = edgeFeature

    faceFeatureReturn.maxEdgeFeature = maxEdgeFeature
    faceFeatureReturn.minEdgeFeature = minEdgeFeature

    # 计算面积平均值和方差
    edge_length = sum(edge_len_list)
    edge_len_average = sum(edge_len_list) / len(edge_len_list)
    edge_len_variance = sum([(m - edge_len_average) ** 2 \
                             for m in edge_len_list]) / len(edge_len_list)

    faceFeatureReturn.perimeter = edge_length
    faceFeatureReturn.edgeLengthAverage = edge_len_average
    faceFeatureReturn.edgeLengthVariance = edge_len_variance

    # 获取面的类型、法线坐标、法线方向
    kind_location_axis = get_face_kind_location_axis(_face)
    faceFeatureReturn.kind = kind_location_axis[0]
    location = kind_location_axis[1]
    if isinstance(location, gp_Pnt):
        faceFeatureReturn.axisLocationX = location.X()
        faceFeatureReturn.axisLocationY = location.Y()
        faceFeatureReturn.axisLocationZ = location.Z()
    axis = kind_location_axis[2]
    if isinstance(axis, gp_Dir):
        faceFeatureReturn.axisDirectX = axis.X()
        faceFeatureReturn.axisDirectY = axis.Y()
        faceFeatureReturn.axisDirectZ = axis.Z()

    return faceFeatureReturn


def get_face_node_feature(_face: TopoDS_Face):
    props = GProp_GProps()
    brepgprop_SurfaceProperties(_face, props)

    faceFeatureReturn = FaceNodeFeature()
    faceFeatureReturn.topoDS_face = _face
    # 面积
    faceFeatureReturn.mass = props.Mass()
    centreOfMass = props.CentreOfMass()
    # 重心坐标
    faceFeatureReturn.centreOfMassX = centreOfMass.X()
    faceFeatureReturn.centreOfMassY = centreOfMass.Y()
    faceFeatureReturn.centreOfMassZ = centreOfMass.Z()

    edge_count = EdgeUtil.get_edge_count_of_face(_face)

    faceFeatureReturn.edgeCount = edge_count

    # 获取面的类型、法线坐标、法线方向
    kind_location_axis = get_face_kind_location_axis(_face)
    faceFeatureReturn.kind = kind_location_axis[0]
    location = kind_location_axis[1]
    if isinstance(location, gp_Pnt):
        faceFeatureReturn.axisLocationX = location.X()
        faceFeatureReturn.axisLocationY = location.Y()
        faceFeatureReturn.axisLocationZ = location.Z()
    axis = kind_location_axis[2]
    if isinstance(axis, gp_Dir):
        faceFeatureReturn.axisDirectX = axis.X()
        faceFeatureReturn.axisDirectY = axis.Y()
        faceFeatureReturn.axisDirectZ = axis.Z()

    return faceFeatureReturn


def get_face_kind_location_axis(topods_face):
    """returns True if the TopoDS_Face is a planar surface"""
    if not isinstance(topods_face, TopoDS_Face):
        return "Not a face", None, None
    surf = BRepAdaptor_Surface(topods_face, True)
    surf_type = surf.GetType()
    if surf_type == GeomAbs_Plane:
        kind = "Plane"
        # look for the properties of the plane
        # first get the related gp_Pln
        gp_pln = surf.Plane()
        location = gp_pln.Location()  # a point of the plane
        normal = gp_pln.Axis().Direction()  # the plane normal
        tuple_to_return = (kind, location, normal)
    elif surf_type == GeomAbs_Cylinder:
        kind = "Cylinder"
        # look for the properties of the cylinder
        # first get the related gp_Cyl
        gp_cyl = surf.Cylinder()
        location = gp_cyl.Location()  # a point of the axis
        axis = gp_cyl.Axis().Direction()  # the cylinder axis
        # then export location and normal to the console output
        tuple_to_return = (kind, location, axis)
    elif surf_type == GeomAbs_Cone:
        kind = "Cone"
        tuple_to_return = (kind, None, None)
    elif surf_type == GeomAbs_Sphere:
        kind = "Sphere"
        tuple_to_return = (kind, None, None)
    elif surf_type == GeomAbs_Torus:
        kind = "Torus"
        tuple_to_return = (kind, None, None)
    elif surf_type == GeomAbs_BezierSurface:
        kind = "Bezier"
        tuple_to_return = (kind, None, None)
    elif surf_type == GeomAbs_BSplineSurface:
        kind = "BSpline"
        tuple_to_return = (kind, None, None)
    elif surf_type == GeomAbs_SurfaceOfRevolution:
        kind = "Revolution"
        tuple_to_return = (kind, None, None)
    elif surf_type == GeomAbs_SurfaceOfExtrusion:
        kind = "Extrusion"
        tuple_to_return = (kind, None, None)
    elif surf_type == GeomAbs_OffsetSurface:
        kind = "Offset"
        tuple_to_return = (kind, None, None)
    elif surf_type == GeomAbs_OtherSurface:
        kind = "Other"
        tuple_to_return = (kind, None, None)
    else:
        tuple_to_return = ("Unknown", None, None)

    return tuple_to_return


# 获取体的所有面
def get_faces_from_solid(_solid: TopoDS_Shape):
    t = TopologyExplorer(_solid)
    faceList = []
    for face in t.faces():
        faceList.append(face)
    return faceList


def get_face_node_from_solid(_solid: TopoDS_Shape, solid_name: str):
    t = TopologyExplorer(_solid)
    faceList = []
    index = 0
    for face in t.faces():
        index = index + 1
        current_face_node = FaceNode(solidName=solid_name,
                                     index=index,
                                     topoDS_face=face,
                                     faceNodeFeature=get_face_node_feature(face))
        faceList.append(current_face_node)
    return faceList


# 获取最大面的面特性
def find_max_face_feature(face_list):
    max_face = None
    max_area = 0.0
    for face in face_list:
        if not isinstance(face, TopoDS_Face):
            continue
        currentArea = get_face_props(face).mass
        if currentArea > max_area:
            max_area = currentArea
            max_face = face
    return max_face


# 获取最小面的面特性
def find_min_face_feature(face_list):
    min_face = None
    min_area = math.inf
    for face in face_list:
        if not isinstance(face, TopoDS_Face):
            continue
        currentArea = get_face_props(face).mass
        if currentArea > min_area:
            min_area = currentArea
            min_face = face
    return min_face


# 获取体的面个数
def get_face_count_of_solid(_solid: TopoDS_Shape):
    t = TopologyExplorer(_solid)
    faceCount = t.number_of_faces()
    return faceCount
