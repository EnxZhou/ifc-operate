import openpyxl


# 零件节点
# class PartNode:
#     def __init__(self,
#                  id,
#                  guid,
#                  node_class,
#                  solid_mass,
#                  centre_of_mass_x,
#                  centre_of_mass_y,
#                  centre_of_mass_z,
#                  surface_area,
#                  max_face_mass,
#                  max_face_centre_of_mass_x,
#                  max_face_centre_of_mass_y,
#                  max_face_centre_of_mass_z,
#                  max_face_axis_location_x,
#                  max_face_axis_location_y,
#                  max_face_axis_location_z,
#                  max_face_axis_direct_x,
#                  max_face_axis_direct_y,
#                  max_face_axis_direct_z,
#                  max_face_perimeter,
#                  max_face_max_edge_centre_x,
#                  max_face_max_edge_centre_y,
#                  max_face_max_edge_centre_z,
#                  max_face_min_edge_centre_x,
#                  max_face_min_edge_centre_y,
#                  max_face_min_edge_centre_z,
#                  max_face_edge_length_average,
#                  max_face_edge_length_variance,
#                  min_face_mass,
#                  min_face_centre_of_mass_x,
#                  min_face_centre_of_mass_y,
#                  min_face_centre_of_mass_z,
#                  min_face_axis_location_x,
#                  min_face_axis_location_y,
#                  min_face_axis_location_z,
#                  min_face_axis_direct_x,
#                  min_face_axis_direct_y,
#                  min_face_axis_direct_z,
#                  min_face_perimeter,
#                  min_face_max_edge_centre_x,
#                  min_face_max_edge_centre_y,
#                  min_face_max_edge_centre_z,
#                  min_face_min_edge_centre_x,
#                  min_face_min_edge_centre_y,
#                  min_face_min_edge_centre_z,
#                  min_face_edge_length_average,
#                  min_face_edge_length_variance,
#                  face_mass_average,
#                  face_mass_variance,
#                  edge_count,
#                  edge_len_sum):
#         self.id = id
#         self.guid = guid
#         self.node_class = node_class
#         self.solid_mass = solid_mass
#         self.centre_of_mass_x = centre_of_mass_x
#         self.centre_of_mass_y = centre_of_mass_y
#         self.centre_of_mass_z = centre_of_mass_z
#         self.surface_area = surface_area
#         self.max_face_mass = max_face_mass
#         self.max_face_centre_of_mass_x = max_face_centre_of_mass_x
#         self.max_face_centre_of_mass_y = max_face_centre_of_mass_y
#         self.max_face_centre_of_mass_z = max_face_centre_of_mass_z
#         self.max_face_axis_location_x = max_face_axis_location_x
#         self.max_face_axis_location_y = max_face_axis_location_y
#         self.max_face_axis_location_z = max_face_axis_location_z
#         self.max_face_axis_direct_x = max_face_axis_direct_x
#         self.max_face_axis_direct_y = max_face_axis_direct_y
#         self.max_face_axis_direct_z = max_face_axis_direct_z
#         self.max_face_perimeter = max_face_perimeter
#         self.max_face_max_edge_centre_x = max_face_max_edge_centre_x
#         self.max_face_max_edge_centre_y = max_face_max_edge_centre_y
#         self.max_face_max_edge_centre_z = max_face_max_edge_centre_z
#         self.max_face_min_edge_centre_x = max_face_min_edge_centre_x
#         self.max_face_min_edge_centre_y = max_face_min_edge_centre_y
#         self.max_face_min_edge_centre_z = max_face_min_edge_centre_z
#         self.max_face_edge_length_average = max_face_edge_length_average
#         self.max_face_edge_length_variance = max_face_edge_length_variance
#         self.min_face_mass = min_face_mass
#         self.min_face_centre_of_mass_x = min_face_centre_of_mass_x
#         self.min_face_centre_of_mass_y = min_face_centre_of_mass_y
#         self.min_face_centre_of_mass_z = min_face_centre_of_mass_z
#         self.min_face_axis_location_x = min_face_axis_location_x
#         self.min_face_axis_location_y = min_face_axis_location_y
#         self.min_face_axis_location_z = min_face_axis_location_z
#         self.min_face_axis_direct_x = min_face_axis_direct_x
#         self.min_face_axis_direct_y = min_face_axis_direct_y
#         self.min_face_axis_direct_z = min_face_axis_direct_z
#         self.min_face_perimeter = min_face_perimeter
#         self.min_face_max_edge_centre_x = min_face_max_edge_centre_x
#         self.min_face_max_edge_centre_y = min_face_max_edge_centre_y
#         self.min_face_max_edge_centre_z = min_face_max_edge_centre_z
#         self.min_face_min_edge_centre_x = min_face_min_edge_centre_x
#         self.min_face_min_edge_centre_y = min_face_min_edge_centre_y
#         self.min_face_min_edge_centre_z = min_face_min_edge_centre_z
#         self.min_face_edge_length_average = min_face_edge_length_average
#         self.min_face_edge_length_variance = min_face_edge_length_variance
#         self.face_mass_average = face_mass_average
#         self.face_mass_variance = face_mass_variance
#         self.edge_count = edge_count
#         self.edge_len_sum = edge_len_sum
#
#
# def get_cols(ws, header_row):
#     cols = {
#         'id': None, 'guid': None, 'class': None,
#         'centreOfMassX': None, 'centreOfMassY': None, 'centreOfMassZ': None,
#         'surface_area':None, 'max_face_mass':None,
#         'max_face_centre_of_mass_x':None, 'max_face_centre_of_mass_y':None, 'max_face_centre_of_mass_z':None,
#         'max_face_axis_location_x':None, 'max_face_axis_location_y':None, 'max_face_axis_location_z':None,
#         'max_face_axis_direct_x':None, 'max_face_axis_direct_y':None, 'max_face_axis_direct_z':None,
#         'max_face_perimeter':None,
#         'max_face_max_edge_centre_x':None, 'max_face_max_edge_centre_y':None, 'max_face_max_edge_centre_z':None,
#         'max_face_min_edge_centre_x':None, 'max_face_min_edge_centre_y':None, 'max_face_min_edge_centre_z':None,
#         'max_face_edge_length_average':None, 'max_face_edge_length_variance':None, 'min_face_mass':None,
#         'min_face_centre_of_mass_x':None, 'min_face_centre_of_mass_y':None, 'min_face_centre_of_mass_z':None,
#         'min_face_axis_location_x':None, 'min_face_axis_location_y':None, 'min_face_axis_location_z':None,
#         'min_face_axis_direct_x':None, 'min_face_axis_direct_y':None, 'min_face_axis_direct_z':None,
#         'min_face_perimeter':None,
#         'min_face_max_edge_centre_x':None, 'min_face_max_edge_centre_y':None, 'min_face_max_edge_centre_z':None,
#         'min_face_min_edge_centre_x':None, 'min_face_min_edge_centre_y':None, 'min_face_min_edge_centre_z':None,
#         'min_face_edge_length_average':None, 'min_face_edge_length_variance':None, 'face_mass_average':None,
#         'face_mass_variance':None, 'edge_count':None,
#         'edgeLenSum': None
#     }
#     header = list(ws[header_row])
#     for index, cell in enumerate(header, start=1):
#         value = cell.value
#         if value in cols:
#             cols[value] = index
#     return cols

class PartNode:
    def __init__(self, **kwargs):
        self.__dict__.update(kwargs)


def get_cols(ws, header_row):
    cols = {}
    header = list(ws[header_row])
    for index, cell in enumerate(header, start=1):
        value = cell.value
        if value:
            cols[value] = index
    return cols


def to_float_or_return(value):
    """
    如果输入是整数或浮点数，则将其转换为浮点数。
    如果输入不是数字类型，则直接返回输入值。

    参数:
    value -- 输入值

    返回:
    转换后的浮点数，或者原始输入值（如果它不是数字类型）
    """
    if isinstance(value, (int, float)):
        return float(value)
    else:
        return value


# 读取生成的solid参数xlsx文件，形成SolidFeature列表
def read_part_node_from_xlsx(file_name):
    wb = openpyxl.load_workbook(file_name, read_only=True)
    ws = wb['solid']
    header_row = 1
    cols = get_cols(ws, header_row)

    part_nodes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        part_node_data = {}
        for attribute, col_index in cols.items():
            try:
                part_node_data[attribute] = to_float_or_return(row[col_index - 1]) if col_index is not None else None
            except IndexError as e:
                print(f"IndexError: Attribute '{attribute}' has an index out of range.")
                print("Row:", row)
                print("Columns:", cols)
                raise e
        part_node = PartNode(**part_node_data)  # Assuming PartNode class exists
        part_nodes.append(part_node)

    return part_nodes
