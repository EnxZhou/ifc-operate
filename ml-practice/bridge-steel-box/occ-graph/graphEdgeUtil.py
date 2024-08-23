import openpyxl


class EdgeOfFaceNode:
    node1 = ""
    node2 = ""

    def __init__(self, node1: str, node2: str):
        self.node1 = node1
        self.node2 = node2


# 读取判断好dist的文件，转换为图结构的edge
def dist_edge_from_xlsx(file_name):
    # 读取xlsx文件中名为“dist”的sheet表
    wb = openpyxl.load_workbook(file_name, read_only=True)
    ws = wb['dist']

    # 获取字段名所在的行号
    header_row = 1
    header = []
    for cell in ws[header_row]:
        header.append(cell.value)

    # 获取每个字段对应的列号
    node1_col = header.index('node1')
    index1_col = header.index('index1')
    node2_col = header.index('node2')
    index2_col = header.index('index2')

    # 遍历sheet表中的数据，并将其写入SolidNode对象
    edge_nodes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        node1 = row[node1_col] + '_' + str(row[index1_col])
        node2 = row[node2_col] + '_' + str(row[index2_col])
        edge_node = EdgeOfFaceNode(node1, node2)
        edge_nodes.append(edge_node)
    return edge_nodes

# 读取节点edge文件，形成图的边数据
def read_edge_from_xlsx(file_name):
    # 读取xlsx文件中名为“edge”的sheet表
    wb = openpyxl.load_workbook(file_name, read_only=True)
    ws = wb['edge']

    # 获取字段名所在的行号
    header_row = 1
    header = []
    for cell in ws[header_row]:
        header.append(cell.value)

    # 获取每个字段对应的列号
    node1_col = header.index('Src')
    node2_col = header.index('Dst')

    # 遍历sheet表中的数据，并将其写入SolidNode对象
    edge_nodes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        node1 = row[node1_col]
        node2 = row[node2_col]
        edge_node = EdgeOfFaceNode(node1, node2)
        edge_nodes.append(edge_node)
    return edge_nodes